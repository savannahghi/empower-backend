"""Patient related tasks."""
import textwrap
import uuid
from typing import Dict

from django.conf import settings
from openpyxl import load_workbook

from sil_advantage.common.models.common_models import Person
from sil_advantage.common.tasks import BaseTaskWithRetry, celery_app
from sil_advantage.notifications.sms.utils import send_custom_sms
from sil_advantage.patients.utils import upload_patient_records
from sil_advantage.segments.models import SegmentMemberAdditionAttributes
from sil_advantage.segments.models.segments import SegmentUpload


@celery_app.task(base=BaseTaskWithRetry)
def send_patient_registration_sms(patient_id: uuid.UUID, sms_intention: str) -> None:
    """Send a patient registration SMS."""
    # Import's here due to circular dependency issues
    from sil_advantage.patients.models import Patient

    patient = Patient.objects.select_related("person", "organisation").get(
        pk=patient_id
    )
    org = patient.organisation
    branch_id = patient.branch_id
    phone_number = patient.person.phone_number
    priority = settings.CELERY_TASK_LOW_PRIORITY

    if phone_number:
        viable_sms_intentions: dict[str, dict] = {
            "PATIENT_REGISTRATION": {
                "patient_id": patient.patient_id,
                "department_id": patient.department_id,
                "workstation_id": patient.workstation_id,
            },
        }
        kwargs = viable_sms_intentions[sms_intention]
        send_custom_sms(
            sms_intention=sms_intention,
            phone_number=phone_number,
            org=org,
            branch_id=branch_id,
            person=patient.person,
            priority=priority,
            **kwargs,
        )


@celery_app.task(base=BaseTaskWithRetry)
def send_patient_communication_consent_otp_sms(
    person_id: uuid.UUID, sms_intention: str, otp_code: str
) -> None:
    """Send a patient communication consent SMS."""
    person = Person.objects.get(id=person_id)
    org = person.organisation
    branch_id = person.branch_id

    phone_number = person.phone_number
    priority = settings.CELERY_TASK_MEDIUM_PRIORITY

    extra_fields = {
        "code": otp_code,
        "workstation_id": person.workstation_id,
    }

    if phone_number:
        send_custom_sms(
            sms_intention,
            phone_number,
            org,
            branch_id,
            person,
            priority,
            **extra_fields,
        )


@celery_app.task(base=BaseTaskWithRetry)
def send_health_id_registration_sms(patient_id: uuid.UUID, sms_intention: str) -> None:
    """Send a health ID once CRM has assigned a Health ID."""
    # Import's here due to circular dependency issues
    from sil_advantage.patients.models import Patient

    patient = Patient.objects.select_related("person", "organisation").get(
        pk=patient_id
    )
    org = patient.organisation
    branch_id = patient.branch_id
    phone_number = patient.person.phone_number
    priority = settings.CELERY_TASK_LOW_PRIORITY
    global_health_id = patient.global_health_id

    if phone_number and global_health_id:
        health_id = " ".join(
            textwrap.wrap(global_health_id, width=len(global_health_id) // 4)
        )
        viable_sms_intentions: dict[str, dict] = {
            "PATIENT_GLOBAL_HEALTH_ID": {"health_id": health_id},
        }
        kwargs = viable_sms_intentions[sms_intention]
        send_custom_sms(
            sms_intention=sms_intention,
            phone_number=phone_number,
            org=org,
            branch_id=branch_id,
            person=patient.person,
            priority=priority,
            **kwargs,
        )


@celery_app.task()
def process_patient_list_upload(
    patient_list_upload_id: uuid.UUID,
    segment_upload_id: uuid.UUID | None = None,
) -> None:
    """Upload patients from a patient list.

    User mapping format:
        {
            "first_name": "<column_header>",
            "last_name": "<column_header>",
            "full_name": "<column_header>",
            "other_names": "<column_header>",
            "age": "<column_header>"
            "gender": "<column_header>"
            "date_of_birth": "<column_header>"
            "phone_number": "<column_header>"
        }

    Args:
        patient_list_upload_id: A `PatientListUpload` instance ID.
        segment_upload_id: A `Segment Upload` instance ID.
    """
    # Import's here due to circular dependency issues
    from sil_advantage.patients.models import PatientListUpload
    from sil_advantage.segments.models import SegmentMember

    patient_list_upload = PatientListUpload.objects.get(id=patient_list_upload_id)
    work_book = load_workbook(patient_list_upload.upload_file.data)
    work_sheet = work_book.active
    resolved_mappings: Dict[str, int] = {}  # {"first_name":0,"last_name":1,...}

    data_rows = []
    for row in work_sheet.iter_rows(values_only=True):
        # only append rows with data
        if not all(cell is None for cell in row):
            data_rows.append(row)

    header_row = data_rows[0]
    patient_rows: list[tuple] = data_rows[1:]

    headers = [str(header).strip() for header in header_row]
    # resolve mapped fields
    for key, value in patient_list_upload.mapping.items():
        if value in headers:
            resolved_mappings[key] = headers.index(value)

    extra_headers = list(set(headers) - set(patient_list_upload.mapping.values()))
    formatted_extra_headers = []

    # resolve unmapped fields
    for header in extra_headers:
        if header == "None":
            continue

        formatted_header = header.lower().strip().replace(" ", "_")
        formatted_extra_headers.append(formatted_header)

        resolved_mappings[formatted_header] = headers.index(header)

    uploaded_persons = upload_patient_records(
        headers=headers,
        patients=patient_rows,
        resolved_mappings=resolved_mappings,
        patient_list_upload=patient_list_upload,
    )

    if patient_list_upload.upload_type == "SEGMENT" and segment_upload_id:
        upload = SegmentUpload.objects.get(id=segment_upload_id)
        segment_member_list = []
        for person in uploaded_persons:
            segment_member = SegmentMember(
                organisation=patient_list_upload.organisation,
                branch_id=patient_list_upload.branch_id,
                cluster_id=patient_list_upload.cluster_id,
                department_id=patient_list_upload.department_id,
                workstation_id=patient_list_upload.workstation_id,
                person=person["person"],
                segment=upload.segment,
                status="CONFIRMED",
                extra_data=person["extras"],
                source=SegmentMemberAdditionAttributes.UPLOAD,
                created_by=patient_list_upload.created_by,
                updated_by=patient_list_upload.created_by,
            )
            segment_member_list.append(segment_member)
        SegmentMember.objects.bulk_create(segment_member_list, ignore_conflicts=True)

        if formatted_extra_headers:
            upload.extra_headers = formatted_extra_headers
            upload.save(update_fields=["extra_headers"])
